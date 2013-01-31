#!/usr/bin/env python
"""

This file reads from the input .xlsx file and writes output files as
LaTeX and .pdf.

Initializing works in a few steps:
   Try importing openpyxl. If this fails, pip install openpyxl.
   Try opening ../schedule_input.xlsx.
   Assert that all needed sheets are available.
   Try opening ../recruit_schedule_template.tex.

31 Jan 13 - For some terrible reason I can't figure out, openpyxl occasionally
loads a bunch of null-valued cells on Macs. Added _unpack() to fix this.

v1.0 - Peter Winter
v2.0 - Patrick Fuller, patrickfuller@gmail.com, 22 Oct 12

"""

import re

# Load appropriate library and file on import
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.style import Color, Fill
wb = load_workbook(filename="../schedule_input.xlsx")

# If a sheet doesn't exist, openpyxl returns null. That's dumb. Throw an error.
sheets = ["Recruit Preferences", "Professor Availability",
          "Manual Overrides", "Travel Weights"]
if not set(sheets) <= set(wb.get_sheet_names()):
    raise Exception(("\n\nDon't rename the spreadsheet tabs!\n\n"
                     "The script expects:\n%s\n" % sheets))

# Also, load tex file here so the code will error if something's amiss
tex_template = open("../recruit_schedule_template.tex").read()


def get_recruit_preferences():
    """
    Gets the professor preferences of each recruit from the .xlsx document.
    Returns [{ name: "", preferences: [] }]
    """
    sheet = wb.get_sheet_by_name("Recruit Preferences")
    rows = _unpack(sheet)
    return [{"name": "%s %s" % (row[1], row[0]), "preferences": row[2:]}
            for row in rows[1:]]


def get_professor_availability():
    """
    Gets the availabilities of professors from the .xlsx document
    Returns [{ name: "", cluster: "", is_available: [] }]
    """
    sheet = wb.get_sheet_by_name("Professor Availability")
    rows = _unpack(sheet)
    return [{"name": row[0],
             "cluster": row[1],
             "is_available": [r in ["Y", "y"] for r in row[2:]]}
            for row in rows[1:]]


def get_manual_overrides():
    """
    Gets unflexible (recruit, professor, slot) times to add as constraints.
    Returns [{ "recruit": "", "professor": "", "slot": # }]
    """
    sheet = wb.get_sheet_by_name("Manual Overrides")
    rows = _unpack(sheet)
    return [{"recruit": "%s %s" % (row[1], row[0]),
             "professor": row[2],
             "slot": int(row[3])}
            for row in rows[1:]]


def get_travel_weights():
    """
    Gets the travel weight matrix from the .xlsx document
    Returns a nested dict, so d[cluster_1][cluster_2] = weight
    """

    sheet = wb.get_sheet_by_name("Travel Weights")
    rows = _unpack(sheet)

    # Gets clusters from the top row of the sheet
    columns = rows[0][1:]

    # Spelling this out, as a nested dictionary comprehension looks meh and
    # this isn't exactly a speed bottleneck
    weights = {}
    for row in rows[1:]:
        cluster = row[0]
        costs = [int(cell) for cell in row[1:]]
        weights[cluster] = {col: cost for col, cost in zip(columns, costs)}
    return weights


def read_schedule_xlsx(filepath):
    """Reads a .xlsx spreadsheet (as defined below) into an object."""
    wb = load_workbook(filename=filepath)
    sheet = wb.get_sheet_by_name("Recruit Schedule")
    rows = _unpack(sheet)
    return [{"name": r[0], "slots": r[1:]} for r in rows[1:]]


def _unpack(sheet):
    """Removes cells with None values and unpacks Cell objects.

    For some reason, openpyxl's garbage_collect() function doesn't always
    work. This seems more reliable.
    """

    # Get highest column. The aptly named sheet.get_highest_row() and
    # sheet.get_highest_column() functions straight up fails on Macs.
    max_column = 0
    for row in sheet.rows:
        column_index = -1
        while not row[column_index].value:
            column_index -= 1
        column_index = len(row) + column_index + 1
        if max_column < column_index:
            max_column = column_index

    return [[r.value for r in row[:max_column]] for row in sheet.rows]


def write_schedule_xlsx(professors, recruits, filepath):
    """
    Writes and saves a .xlsx spreadsheet containing three tabs - Professor
    Schedule, Recruit Schedule, and Recruit Clusters. The first two should be
    self explanatory, and the third helps for assigning runners.
    """

    # Make a workbook object and a row-1 header
    wb = Workbook()
    num_slots = len(professors[0]["slots"])
    header = [""] + ["Slot %d" % (s + 1) for s in range(num_slots)]

    # Make a sheet for the professor schedule
    ws = wb.create_sheet(index=0, title="Professor Schedule")
    ws.append(header)
    for professor in professors:
        ws.append([professor["name"]] + professor["slots"])

    # Make another sheet for the recruit schedule
    ws = wb.create_sheet(index=0, title="Recruit Schedule")
    ws.append(header)
    for recruit in recruits:
        ws.append([recruit["name"]] + recruit["slots"])

    # Color cells according to whether or not it fulfills a recruit request
    for row in ws.rows[1:]:
        for recruit in recruits:
            if recruit["name"] == row[0].value:
                break
        for cell in row[1:]:
            cell.style.fill.fill_type = Fill.FILL_SOLID

            # These colors are hideous, but I'm sick of openpyxl and don't want
            # to figure out its arbitrary Color constructor
            if not cell.value:
                color = Color.DARKYELLOW
            elif cell.value in recruit["preferences"]:
                color = Color.DARKGREEN
            else:
                color = Color.DARKRED

            cell.style.fill.start_color.index = color

    # Make a final sheet showing recruit clusters (helps w/ assigning runners)
    ws = wb.create_sheet(index=0, title="Recruit Clusters")
    ws.append(header)
    for recruit in recruits:
        recruit["clusters"] = ["Poster"] * len(recruit["slots"])
        for i, slot in enumerate(recruit["slots"]):
            for professor in professors:
                if professor["name"] == slot:
                    recruit["clusters"][i] = professor["cluster"]
                    break
        ws.append([recruit["name"]] + recruit["clusters"])
    wb.save(filepath)


def write_tex_file(recruit, filepath):
    """Writes a .tex file containing a recruit's schedule."""

    # Copy the template object so we don't have to deal with re-reading files
    tex = tex_template[:]

    # Define replacement logic with a closure
    def tex_repl(match):
        """Uses python.re's MatchObject to parse patterns."""
        flag = match.group()[2:-1]

        if flag == "name":
            return recruit["name"]

        # If it's a slot, extract the number and use it to find a professor
        flag, num = flag.split("_")
        num = int(num) - 1
        if flag == "slot":
            prof = recruit["slots"][num]
            return "Meet with Professor " + prof if prof else "Poster session"

        # Exceptions for the assholes
        raise Exception("A given #(*) .tex flag is not supported!")

    # Iterate through and replace flags
    tex = re.sub("#\(\w+\)", tex_repl, tex)

    with open(filepath, 'w') as outfile:
        outfile.write(tex)

if __name__ == "__main__":
    print get_recruit_preferences()
    print get_professor_availability()
    print get_manual_overrides()
    print get_travel_weights()
